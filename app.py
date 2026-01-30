import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import io
import os
import re
from PIL import Image, ImageOps

# 匯入我們拆分出去的模組 (必須確保 styles.py 與 utils.py 在同一資料夾)
from styles import inject_custom_styles
from utils import (
    format_date_roc, format_layout, safe_float_convert, 
    crop_image_to_ratio, calculate_cell_pixels, parse_transcript_pdf
)

# --- 1. 頁面設定 (★請在這裡修改 App 名稱與圖示★) ---
st.set_page_config(
    page_title="studio",      # <-- 已修改：瀏覽器標籤名稱
    page_icon="https://raw.githubusercontent.com/vary530/survey_app/main/my_logo.png", # <-- 將網址填入 page_icon 才能正確顯示
    layout="centered",
    initial_sidebar_state="collapsed"
)

# --- 核心邏輯 ---
TEMPLATE_FILE = "template.xlsx"
MAIN_ORDER = [
    "物件類型", "案名", "地址", "社區名稱", 
    "地上層", "地下層", "位於樓層", "格局", 
    "售價", "登記總建坪", "主建物坪數", "附屬建坪數", "公設坪數", "不含車位坪數", 
    "車位坪數", "車位形式", "車位樓層", "汽車編號", "機車位樓層", "機車編號", 
    "使用現況", "總戶數", "同層戶數", "電梯數", "有無警衛", "管理費", "繳納方式", 
    "建築完成日", "瓦斯", "學校", "市場", "公園", "公設比", 
    "建物KEY", "座向", "土地面積", "權利範圍", 
    "冒泡位置圖", "承辦人電話", "委託契約書編號" 
]
OTHER_ORDER = [
    "房地合一", "面道路", "貸款設定", "車位價格", "房屋單價"
]

# --- 2. 注入視覺設計 ---
inject_custom_styles()

def main():
    if not os.path.exists(TEMPLATE_FILE):
        st.error(f"系統錯誤：找不到 {TEMPLATE_FILE}")
        return
    try:
        wb = load_workbook(TEMPLATE_FILE)
        target_sheet = None
        for sheetname in wb.sheetnames:
            if "物調表" in sheetname:
                target_sheet = wb[sheetname]
                break
        if target_sheet is None: target_sheet = wb.active 
    except Exception as e:
        st.error(f"系統錯誤：讀取模板失敗 {e}")
        return

    # 掃描 Excel 模板標籤
    label_to_coord = {}
    scanned_items = []
    for row in target_sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and '"""' in cell.value:
                raw_txt = cell.value
                label_name = ""
                content_part = ""
                
                match_star = re.search(r'\*(.*?)\*(.*)', raw_txt)
                if match_star:
                    label_name = match_star.group(1).strip()
                    content_part = match_star.group(2).replace('"""', '')
                else:
                    label_name = raw_txt.replace('"""', '').strip()
                    content_part = label_name

                options = []
                input_type = "text"
                
                if "□" in content_part:
                    input_type = "select"
                    segments = content_part.split('□')
                    options = [s.strip() for s in segments if s.strip()]
                    options.insert(0, "請選擇...")
                
                if "特色" in label_name or "說明" in label_name:
                    input_type = "textarea"
                elif "冒泡" in label_name:
                    input_type = "image_upload"

                item_data = {
                    "label": label_name,
                    "coordinate": cell.coordinate,
                    "type": input_type,
                    "options": options
                }
                label_to_coord[label_name] = cell.coordinate
                scanned_items.append(item_data)

    # --- 介面呈現 ---
    st.markdown("<h1>studio</h1>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>YUNGYI PROPERTY INTEGRATION</div>", unsafe_allow_html=True)

    st.markdown("<div style='color:#c5a065; font-size:15px; font-weight:bold; margin-bottom:10px; margin-top:20px;'>智慧匯入中心</div>", unsafe_allow_html=True)
    
    uploaded_pdf = st.file_uploader("點此上傳建物謄本 (PDF)", type=['pdf'])
    
    # PDF 處理邏輯
    if uploaded_pdf:
        if 'last_uploaded_pdf' not in st.session_state or st.session_state.last_uploaded_pdf != uploaded_pdf.name:
            with st.spinner("分析中..."):
                parsed = parse_transcript_pdf(uploaded_pdf)
                st.session_state.pdf_parsed_data = parsed
                st.session_state.last_uploaded_pdf = uploaded_pdf.name
        
        if 'pdf_parsed_data' in st.session_state:
            data = st.session_state.pdf_parsed_data
            
            grid_html = f"""
            <div class="dashboard-grid">
                <div class="dash-item"><div class="dash-label">地址</div><div class="dash-value">{data.get('地址', '-')}</div></div>
                <div class="dash-item"><div class="dash-label">建築完成日</div><div class="dash-value">{data.get('建築完成日', '-')}</div></div>
                <div class="dash-item"><div class="dash-label">主建物坪數</div><div class="dash-value">{data.get('主建物坪數', '-')}</div></div>
                <div class="dash-item"><div class="dash-label">附屬建坪數</div><div class="dash-value">{data.get('附屬建坪數', '-')}</div></div>
                <div class="dash-item"><div class="dash-label">地上層</div><div class="dash-value">{data.get('地上層', '-')}</div></div>
                <div class="dash-item"><div class="dash-label">位於樓層</div><div class="dash-value">{data.get('位於樓層', '-')}</div></div>
            </div>
            """
            st.markdown(grid_html, unsafe_allow_html=True)
            
            if st.button("匯入建物基本資料", type="primary"):
                count = 0
                for pdf_key, pdf_val in st.session_state.pdf_parsed_data.items():
                    target_coord = None
                    if pdf_key in label_to_coord: target_coord = label_to_coord[pdf_key]
                    if not target_coord:
                        for lbl, coord in label_to_coord.items():
                            if pdf_key in lbl or lbl in pdf_key:
                                target_coord = coord
                                break
                    if target_coord:
                        st.session_state[target_coord] = pdf_val
                        count += 1
                if count > 0:
                    st.success("資料已匯入")

    user_inputs = {} 
    uploaded_map_image = None
    scanned_dict = {item["label"]: item for item in scanned_items}

    # 表單渲染
    with st.form("survey_form"):
        st.markdown("<div style='color:#c5a065; font-size:15px; font-weight:bold; margin-bottom:15px;'>不動產基本資料</div>", unsafe_allow_html=True)

        for label in MAIN_ORDER:
            found_key = label if label in scanned_dict else None
            if not found_key:
                for k in scanned_dict.keys():
                    if label in k or k in label:
                        found_key = k
                        break
            
            if found_key:
                item = scanned_dict[found_key]
                coord = item["coordinate"]
                
                if label == "地址":
                    val = st.text_input(label, key=coord)
                    user_inputs[coord] = val
                    if val:
                        map_url = f"https://www.google.com/maps/search/?api=1&query={val}"
                        st.markdown(f"<div style='text-align:right; margin-top:-5px; margin-bottom:10px;'><a href='{map_url}' target='_blank' style='font-size:12px; color:#888; text-decoration:none;'>📍 開啟地圖</a></div>", unsafe_allow_html=True)
                
                elif item["type"] == "select":
                    val = st.selectbox(found_key, item["options"], key=coord)
                    user_inputs[coord] = val if val != "請選擇..." else ""
                
                elif item["type"] == "textarea":
                    val = st.text_area(found_key, key=coord, height=120)
                    user_inputs[coord] = val

                elif item["type"] == "image_upload":
                    st.markdown(f"<div style='margin-top:15px; margin-bottom:5px; font-size:14px; color:#c5a065;'>{found_key}</div>", unsafe_allow_html=True)
                    uploaded_map_image = st.file_uploader("", type=['jpg', 'png', 'jpeg'], key=coord, label_visibility="collapsed")
                    st.markdown("<div style='font-size:12px; color:#666; margin-top:-5px;'>* 圖片將自動「置中剪裁 (27:16)」並拉伸填滿 Excel 儲存格</div>", unsafe_allow_html=True)
                    user_inputs[coord] = ""
                else:
                    placeholder_txt = ""
                    # 提示文字邏輯
                    if "房屋單價" in found_key or "公設比" in found_key:
                        placeholder_txt = "輸入數字0系統匯出自動計算"
                    elif "不含車位坪數" in found_key:
                        placeholder_txt = "輸入數字0系統匯出自動計算"
                    
                    val = st.text_input(found_key, key=coord, placeholder=placeholder_txt)
                    user_inputs[coord] = val
                
                if found_key in scanned_dict: del scanned_dict[found_key]

        if any(k in scanned_dict for k in OTHER_ORDER):
            st.markdown("<hr style='border-color: rgba(255,255,255,0.05); margin: 30px 0;'>", unsafe_allow_html=True)
            for label in OTHER_ORDER:
                if label in scanned_dict:
                    item = scanned_dict[label]
                    coord = item["coordinate"]
                    
                    if item["type"] == "select":
                        val = st.selectbox(label, item["options"], key=coord)
                        user_inputs[coord] = val if val != "請選擇..." else ""
                    elif item["type"] == "textarea":
                        val = st.text_area(label, key=coord, height=100)
                        user_inputs[coord] = val
                    else:
                        placeholder_txt = ""
                        if "房屋單價" in label or "公設比" in label:
                            placeholder_txt = "輸入數字0系統匯出自動計算"
                        elif "不含車位坪數" in label:
                            placeholder_txt = "輸入數字0系統匯出自動計算"
                        
                        val = st.text_input(label, key=coord, placeholder=placeholder_txt)
                        user_inputs[coord] = val
                    
                    del scanned_dict[label]

        if scanned_dict:
            st.markdown("<hr style='border-color: rgba(255,255,255,0.05); margin: 30px 0;'>", unsafe_allow_html=True)
            for label, item in scanned_dict.items():
                coord = item["coordinate"]
                if item["type"] == "select":
                    val = st.selectbox(label, item["options"], key=coord)
                    user_inputs[coord] = val if val != "請選擇..." else ""
                elif item["type"] == "textarea":
                    val = st.text_area(label, key=coord, height=100)
                    user_inputs[coord] = val
                else:
                    placeholder_txt = ""
                    if "房屋單價" in label or "公設比" in label:
                        placeholder_txt = "輸入數字0系統匯出自動計算"
                    elif "不含車位坪數" in label:
                        placeholder_txt = "輸入數字0系統匯出自動計算"

                    val = st.text_input(label, key=coord, placeholder=placeholder_txt)
                    user_inputs[coord] = val

        st.markdown("<br>", unsafe_allow_html=True)
        submitted = st.form_submit_button("匯出至Excel")

    if submitted:
        wb_output = load_workbook(TEMPLATE_FILE)
        ws_output = wb_output[target_sheet.title]

        coord_to_header = {item["coordinate"]: item["label"] for item in scanned_items}
        image_coords = [item["coordinate"] for item in scanned_items if item["type"] == "image_upload"]

        # 取得座標供計算
        coord_price = next((k for k, v in coord_to_header.items() if "售價" in v), None)
        coord_total_area = next((k for k, v in coord_to_header.items() if "登記總建坪" in v), None)
        coord_area_no_parking = next((k for k, v in coord_to_header.items() if "不含車位" in v), None)
        coord_parking_area = next((k for k, v in coord_to_header.items() if "車位坪數" in v), None)
        coord_public_area = next((k for k, v in coord_to_header.items() if "公設坪數" in v), None)
        coord_unit_price = next((k for k, v in coord_to_header.items() if "房屋單價" in v), None)
        coord_public_ratio = next((k for k, v in coord_to_header.items() if "公設比" in v), None)
        coord_main_area = next((k for k, v in coord_to_header.items() if "主建物" in v), None)
        coord_annex_area = next((k for k, v in coord_to_header.items() if "附屬" in v), None)

        # 1. 計算不含車位坪數 (主+附+公) - 恢復自動計算
        if coord_area_no_parking and user_inputs.get(coord_area_no_parking) == "0":
            try:
                a_main = safe_float_convert(user_inputs.get(coord_main_area))
                a_annex = safe_float_convert(user_inputs.get(coord_annex_area))
                a_pub = safe_float_convert(user_inputs.get(coord_public_area))
                user_inputs[coord_area_no_parking] = str(round(a_main + a_annex + a_pub, 3))
            except: pass

        # 2. 計算登記總建坪 (已移除自動計算，保留手動輸入)
        
        # 3. 計算房屋單價
        if coord_unit_price and user_inputs.get(coord_unit_price) == "0":
            try:
                p = safe_float_convert(user_inputs.get(coord_price))
                a = safe_float_convert(user_inputs.get(coord_area_no_parking))
                if a > 0:
                    res = round(p / a, 2)
                    user_inputs[coord_unit_price] = str(res)
            except: pass

        # 4. 計算公設比
        if coord_public_ratio and user_inputs.get(coord_public_ratio) == "0":
            try:
                pub = safe_float_convert(user_inputs.get(coord_public_area))
                a = safe_float_convert(user_inputs.get(coord_area_no_parking))
                if a > 0:
                    res = round((pub / a) * 100, 1)
                    user_inputs[coord_public_ratio] = f"{res}%"
            except: pass

        # 寫入 Excel
        for coord, value in user_inputs.items():
            if coord in image_coords:
                continue

            cell = ws_output[coord]
            final_val = value if value else ""
            header = coord_to_header.get(coord, "")
            
            if "完成日" in header or "日期" in header:
                final_val = format_date_roc(final_val)
            elif "格局" in header:
                final_val = format_layout(final_val)
            
            # 自動加萬
            keywords_for_wan = ["售價", "單價", "價格", "貸款"]
            if any(k in header for k in keywords_for_wan) and final_val:
                v_str = str(final_val).strip()
                if v_str.replace('.', '', 1).isdigit() and "萬" not in v_str:
                    final_val = f"{v_str}萬"
            
            # 管理費自動加元
            if "管理費" in header and final_val:
                v_str = str(final_val).strip()
                if v_str and "元" not in v_str:
                    final_val = f"{v_str}元"

            # 土地面積自動加坪
            if "土地面積" in header and final_val:
                v_str = str(final_val).strip()
                if v_str and "坪" not in v_str:
                    final_val = f"{v_str}坪"

            # 戶數自動加戶 (同層戶數, 總戶數)
            if ("同層戶數" in header or "總戶數" in header) and final_val:
                v_str = str(final_val).strip()
                if v_str and "戶" not in v_str:
                    final_val = f"{v_str}戶"

            cell.value = final_val

        # 圖片處理
        if uploaded_map_image:
            try:
                target_map_coord = None
                for item in scanned_items:
                    if "冒泡" in item["label"]:
                        target_map_coord = item["coordinate"]
                        break
                
                if target_map_coord:
                    ws_output[target_map_coord].value = ""

                    pil_img = Image.open(uploaded_map_image)
                    pil_img = ImageOps.exif_transpose(pil_img)
                    cropped_img = crop_image_to_ratio(pil_img, 27, 16)
                    
                    img_byte_arr = io.BytesIO()
                    cropped_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    img = ExcelImage(img_byte_arr)
                    calc_w, calc_h = calculate_cell_pixels(ws_output, target_map_coord)
                    img.width = calc_w
                    img.height = calc_h
                    
                    ws_output.add_image(img, target_map_coord)
            except Exception as e:
                st.warning(f"圖片處理異常: {e}")

        # 下載檔案處理
        id_coord = None
        name_coord = None
        for item in scanned_items:
            if "委託" in item["label"] and "編號" in item["label"]:
                id_coord = item["coordinate"]
            if "案名" in item["label"]:
                name_coord = item["coordinate"]
        
        file_id = user_inputs.get(id_coord, "無編號") if id_coord else "無編號"
        file_name = user_inputs.get(name_coord, "無案名") if name_coord else "無案名"
        safe_filename = f"{file_id}{file_name}.xlsx"
        safe_filename = "".join([c for c in safe_filename if c.isalpha() or c.isdigit() or c in " ._-()[\u4e00-\u9fa5]"])

        output_buffer = io.BytesIO()
        wb_output.save(output_buffer)
        output_buffer.seek(0)

        st.success(f"整合完成 目前已可供下載Excel：{safe_filename}")
        st.markdown("<p style='font-size:12px; color:#888; text-align:center;'>💡 iPhone 用戶：點擊下載後若跳轉至預覽畫面，<br>請按左上角「完成」或瀏覽器「返回」鍵即可回到此頁面。</p>", unsafe_allow_html=True)
        
        st.download_button(
            label="下載Excel檔案",
            data=output_buffer,
            file_name=safe_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__ == "__main__":
    main()