import re
import pdfplumber
import io
from openpyxl.utils import get_column_letter
import streamlit as st

def full_to_half(s):
    if not s: return ""
    return s.translate(str.maketrans('０１２３４５６７８９', '0123456789'))

def chinese_to_arabic(cn_str):
    if not cn_str: return ""
    cn_map = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5, '六': 6, '七': 7, '八': 8, '九': 9, '十': 10, '0':0, '1':1, '2':2, '3':3, '4':4, '5':5, '6':6, '7':7, '8':8, '9':9}
    clean_str = cn_str.replace('層', '').replace('樓', '').strip()
    if clean_str.isdigit(): return str(int(clean_str))
    try:
        val = 0
        if len(clean_str) == 1: val = cn_map.get(clean_str, 0)
        elif len(clean_str) == 2:
            if clean_str[0] == '十': val = 10 + cn_map.get(clean_str[1], 0)
            elif clean_str[1] == '十': val = cn_map.get(clean_str[0], 0) * 10
        elif len(clean_str) == 3:
             val = cn_map.get(clean_str[0], 0) * 10 + cn_map.get(clean_str[2], 0)
        return str(val) if val > 0 else cn_str
    except: return cn_str

def format_date_roc(date_str):
    if not date_str: return ""
    match = re.match(r'(\d+)[/.-](\d+)[/.-](\d+)', date_str)
    if match:
        y, m, d = match.groups()
        return f"民國{y}年{m}月{d}日"
    return date_str

def format_layout(layout_str):
    if not layout_str: return ""
    parts = re.split(r'[/, .]', layout_str)
    parts = [p for p in parts if p.strip()]
    result = ""
    if len(parts) >= 1: result += f"{parts[0]}房"
    if len(parts) >= 2: result += f"{parts[1]}廳"
    if len(parts) >= 3: result += f"{parts[2]}衛浴"
    if len(parts) >= 4: result += f"{parts[3]}陽台"
    return result if result else layout_str

def safe_float_convert(value):
    """安全轉換字串為浮點數，失敗回傳 0.0"""
    try:
        if not value: return 0.0
        clean_val = re.sub(r'[^\d.]', '', str(value))
        return float(clean_val)
    except:
        return 0.0

def crop_image_to_ratio(image, target_ratio_w=27, target_ratio_h=16):
    """將圖片置中剪裁為指定長寬比"""
    original_w, original_h = image.size
    target_aspect = target_ratio_w / target_ratio_h
    current_aspect = original_w / original_h

    if current_aspect > target_aspect:
        new_w = int(original_h * target_aspect)
        offset = (original_w - new_w) // 2
        box = (offset, 0, offset + new_w, original_h)
    else:
        new_h = int(original_w / target_aspect)
        offset = (original_h - new_h) // 2
        box = (0, offset, original_w, offset + new_h)
    
    return image.crop(box)

def calculate_cell_pixels(ws, coord):
    """計算 Excel 儲存格 (含合併) 的像素大小"""
    target_range = None
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            target_range = merged_range
            break
    
    if target_range:
        min_col, min_row, max_col, max_row = target_range.min_col, target_range.min_row, target_range.max_col, target_range.max_row
    else:
        c = ws[coord]
        min_col, min_row, max_col, max_row = c.column, c.row, c.column, c.row

    total_width = 0
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        cw = ws.column_dimensions[col_letter].width
        if cw is None: cw = 9 
        total_width += cw * 7.7 
        
    total_height = 0
    for row_idx in range(min_row, max_row + 1):
        rh = ws.row_dimensions[row_idx].height
        if rh is None: rh = 15
        total_height += rh * 1.34 
        
    return total_width, total_height

def parse_transcript_pdf(pdf_file):
    data = {}
    full_text = ""
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                full_text += page.extract_text() + "\n"
        
        lines = full_text.split('\n')
        address_prefix = ""
        address_road = ""
        for i, line in enumerate(lines):
            line = line.strip()
            if "建物標示部" in line:
                for offset in range(1, 5):
                    if i + offset < len(lines):
                        txt = lines[i+offset]
                        match = re.search(r'(.+?[市縣].+?[區鄉鎮市])', txt)
                        if match:
                            address_prefix = match.group(1)
                            break
            if "建物門牌" in line:
                parts = line.split("建物門牌")
                if len(parts) > 1 and parts[1].strip():
                    address_road = parts[1].strip()
                elif i+1 < len(lines):
                    address_road = lines[i+1].strip()

        if address_prefix or address_road:
            full_addr = f"{address_prefix}{address_road}"
            data["地址"] = full_to_half(full_addr).replace(" ", "")

        date_match = re.search(r'建築完成日期\s*([民國\d]+年\d+月\d+日)', full_text)
        if date_match: data["建築完成日"] = date_match.group(1)

        layer_m2_matches = re.findall(r'層次面積\s*([\d\.]+)\s*平方公尺', full_text)
        if layer_m2_matches:
            total_main_m2 = sum(float(x) for x in layer_m2_matches)
            data["主建物坪數"] = str(round(total_main_m2 * 0.3025, 3))

        try:
            start = full_text.find("附屬建物用途")
            end = full_text.find("共有部分")
            if start != -1:
                sub_text = full_text[start:end] if end != -1 else full_text[start:]
                annex_matches = re.findall(r'面積\s*([\d\.]+)\s*平方公尺', sub_text)
                if annex_matches:
                    total_annex_m2 = sum(float(x) for x in annex_matches)
                    data["附屬建坪數"] = str(round(total_annex_m2 * 0.3025, 3))
        except: pass

        floors_match = re.search(r'層數\s*(\d+)層', full_text)
        if floors_match: data["地上層"] = str(int(floors_match.group(1)))

        layer_match = re.search(r'層次\s*([^\d\s]+)層', full_text)
        if layer_match and "面積" not in layer_match.group(0):
            data["位於樓層"] = chinese_to_arabic(layer_match.group(1))
        
    except Exception as e:
        st.error(f"PDF 解析錯誤: {e}")
    return data